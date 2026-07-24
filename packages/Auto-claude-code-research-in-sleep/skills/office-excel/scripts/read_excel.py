#!/usr/bin/env python3
# Copyright Amazon.com, Inc. or its affiliates. All Rights Reserved.
# SPDX-License-Identifier: MIT-0
"""Read an Excel (.xlsx) file and output its content as Markdown tables."""
import sys
import os


def read_excel_to_markdown(xlsx_path: str, sheet_name: str | None = None) -> str:
    """Read an .xlsx file and convert to markdown tables."""
    from openpyxl import load_workbook

    if not os.path.exists(xlsx_path):
        print(f"Error: File not found: {xlsx_path}", file=sys.stderr)
        sys.exit(1)

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    sheets = [sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.sheetnames
    lines = []

    for sname in sheets:
        ws = wb[sname]
        lines.append(f"## Sheet: {sname}")
        lines.append("")

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            lines.append("*(empty sheet)*")
            lines.append("")
            continue

        max_cols = max(len(r) for r in rows)

        for i, row in enumerate(rows):
            cells = []
            for j in range(max_cols):
                val = row[j] if j < len(row) else ""
                cells.append(str(val) if val is not None else "")
            lines.append("| " + " | ".join(cells) + " |")
            if i == 0:
                lines.append("| " + " | ".join(["---"] * max_cols) + " |")

        lines.append("")

    wb.close()
    return "\n".join(lines)


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: read_excel.py <input.xlsx> [sheet_name] [output.md]", file=sys.stderr)
        sys.exit(1)

    xlsx_path = sys.argv[1]
    sheet = sys.argv[2] if len(sys.argv) > 2 and not sys.argv[2].endswith(".md") else None
    output = None
    for arg in sys.argv[2:]:
        if arg.endswith(".md"):
            output = arg

    result = read_excel_to_markdown(xlsx_path, sheet)

    if output:
        with open(output, "w", encoding="utf-8") as f:
            f.write(result)
        print(f"Markdown saved to: {output}")
    else:
        print(result)
