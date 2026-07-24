#!/usr/bin/env python3
# Copyright Amazon.com, Inc. or its affiliates. All Rights Reserved.
# SPDX-License-Identifier: MIT-0
"""Create an Excel (.xlsx) file from Markdown table content or JSON data."""
import sys
import os
import json
import re


def markdown_tables_to_excel(md_path: str, output_path: str):
    """Convert markdown tables to an Excel file. Each table becomes a sheet."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    if not os.path.exists(md_path):
        print(f"Error: File not found: {md_path}", file=sys.stderr)
        sys.exit(1)

    with open(md_path, "r", encoding="utf-8") as f:
        content = f.read()

    wb = Workbook()
    wb.remove(wb.active)

    current_sheet_name = "Sheet1"
    table_rows = []
    sheet_count = 0

    def flush_table():
        nonlocal sheet_count, table_rows, current_sheet_name
        if not table_rows:
            return
        sheet_count += 1
        name = current_sheet_name if current_sheet_name != "Sheet1" or sheet_count == 1 else f"Sheet{sheet_count}"
        ws = wb.create_sheet(title=name[:31])
        for i, row_data in enumerate(table_rows):
            for j, cell_val in enumerate(row_data):
                cell = ws.cell(row=i + 1, column=j + 1, value=cell_val)
                if i == 0:
                    cell.font = Font(bold=True)
        for col in ws.columns:
            max_len = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)
        table_rows = []

    for line in content.split("\n"):
        stripped = line.strip()

        heading_match = re.match(r"^#{1,3}\s+(?:Sheet:\s*)?(.+)$", stripped)
        if heading_match and not table_rows:
            current_sheet_name = heading_match.group(1).strip()
            continue

        if stripped.startswith("|") and stripped.endswith("|"):
            cells = [c.strip() for c in stripped.strip("|").split("|")]
            if all(re.match(r"^[-:]+$", c) for c in cells):
                continue
            table_rows.append(cells)
        elif table_rows:
            flush_table()
            current_sheet_name = "Sheet1"

    flush_table()

    if not wb.sheetnames:
        wb.create_sheet(title="Sheet1")

    wb.save(output_path)
    print(f"Excel file saved to: {output_path}")


def json_to_excel(json_path: str, output_path: str):
    """Convert JSON array of objects to Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    if not isinstance(data, list) or not data:
        print("Error: JSON must be a non-empty array of objects", file=sys.stderr)
        sys.exit(1)

    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    headers = list(data[0].keys())
    for j, h in enumerate(headers):
        cell = ws.cell(row=1, column=j + 1, value=h)
        cell.font = Font(bold=True)

    for i, row_obj in enumerate(data):
        for j, h in enumerate(headers):
            ws.cell(row=i + 2, column=j + 1, value=row_obj.get(h, ""))

    wb.save(output_path)
    print(f"Excel file saved to: {output_path}")


def markdown_string_to_excel(md_content: str, output_path: str):
    """Convert a markdown string (read from stdin) to an Excel file."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    wb.remove(wb.active)

    current_sheet_name = "Sheet1"
    table_rows = []
    sheet_count = 0

    def flush_table():
        nonlocal sheet_count, table_rows, current_sheet_name
        if not table_rows:
            return
        sheet_count += 1
        name = current_sheet_name if current_sheet_name != "Sheet1" or sheet_count == 1 else f"Sheet{sheet_count}"
        ws = wb.create_sheet(title=name[:31])
        for i, row_data in enumerate(table_rows):
            for j, cell_val in enumerate(row_data):
                cell = ws.cell(row=i + 1, column=j + 1, value=cell_val)
                if i == 0:
                    cell.font = Font(bold=True)
        for col in ws.columns:
            max_len = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)
        table_rows = []

    for line in md_content.split("\n"):
        stripped = line.strip()

        heading_match = re.match(r"^#{1,3}\s+(?:Sheet:\s*)?(.+)$", stripped)
        if heading_match and not table_rows:
            current_sheet_name = heading_match.group(1).strip()
            continue

        if stripped.startswith("|") and stripped.endswith("|"):
            cells = [c.strip() for c in stripped.strip("|").split("|")]
            if all(re.match(r"^[-:]+$", c) for c in cells):
                continue
            table_rows.append(cells)
        elif table_rows:
            flush_table()
            current_sheet_name = "Sheet1"

    flush_table()

    if not wb.sheetnames:
        wb.create_sheet(title="Sheet1")

    wb.save(output_path)
    print(f"Excel file saved to: {output_path}")


if __name__ == "__main__":
    # Mode 1: stdin piped content -> xlsx
    # Usage: echo "markdown" | write_excel.py --stdin <output.xlsx>
    if len(sys.argv) >= 3 and sys.argv[1] == "--stdin":
        output_path = sys.argv[2]
        md_content = sys.stdin.read()
        if not md_content.strip():
            print("Error: No content received from stdin", file=sys.stderr)
            sys.exit(1)
        markdown_string_to_excel(md_content, output_path)
    # Mode 2: file -> xlsx
    elif len(sys.argv) >= 3:
        input_path = sys.argv[1]
        output_path = sys.argv[2]
        if input_path.endswith(".json"):
            json_to_excel(input_path, output_path)
        else:
            markdown_tables_to_excel(input_path, output_path)
    else:
        print("Usage:", file=sys.stderr)
        print("  write_excel.py <input.md|input.json> <output.xlsx>", file=sys.stderr)
        print("  echo 'md content' | write_excel.py --stdin <output.xlsx>", file=sys.stderr)
        sys.exit(1)
